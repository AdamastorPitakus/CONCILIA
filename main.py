import pdfplumber
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
import logging

# Configurando o logging para salvar em um arquivo de log com o nome app.log,
# com o nível de log em INFO e o formato da mensagem a aparecer no terminal.
logging.basicConfig(filename='app.log', filemode='w', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.info('Programa iniciado')

def extract_data_from_pdf(pdf_path):
    
    logging.info(f"Extraindo dados do PDF: {pdf_path}")

    data = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            page_data = page_text.split('\n')
            data.extend(page_data)

    logging.info(f"Dados extraídos do PDF: {pdf_path}")
    return data

def criar_planilha_conciliacao_bancaria(pdf_path):
    
    logging.info("Criando planilha de conciliação bancária")

    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação Bancária"
    ws['A1'] = "Data da Transação"
    ws['B1'] = "Descrição/Referência"
    ws['C1'] = "Valor"
    ws['D1'] = "Saldo Disponível"
    ws['E1'] = "Identificador de Transação"
    ws['F1'] = "Categoria/Conta Contábil"
    ws['G1'] = "Status de Conciliação"
    ws['H1'] = "Notas/Comentários"

    logging.info("Planilha de conciliação bancária criada")

    dados_extraidos = extract_data_from_pdf(pdf_path)

    # Inicializar listas vazias para cada coluna
    coluna_data = []
    coluna_descricao = []
    coluna_valor = []
    coluna_saldo_disponivel = []
    coluna_identificador_transacao = []
    coluna_categoria_conta_contabil = []
    coluna_status_conciliacao = []
    coluna_notas_comentarios = []

    # Processar os dados extraídos
    for dado in dados_extraidos:
        if dado.strip():  # Verificar se a linha não está em branco
            if dado[0].isdigit():
                coluna_data.append(dado)
            elif dado[0] == 'R':
                coluna_valor.append(dado)
            elif dado[0] == 'S':
                coluna_saldo_disponivel.append(dado)
            elif dado[0] == 'I':
                coluna_identificador_transacao.append(dado)
            elif dado[0] == 'C':
                coluna_categoria_conta_contabil.append(dado)
            elif dado[0] == 'E':
                coluna_status_conciliacao.append(dado)
            elif dado[0] == 'N':
                coluna_notas_comentarios.append(dado)
            else:
                coluna_descricao.append(dado)

    # Garantir que todas as listas tenham o mesmo comprimento
    min_length = min(len(coluna_data), len(coluna_descricao), len(coluna_valor),
                     len(coluna_saldo_disponivel), len(coluna_identificador_transacao),
                     len(coluna_categoria_conta_contabil), len(coluna_status_conciliacao),
                     len(coluna_notas_comentarios))

    # Percorrer e gravar os dados nas colunas da planilha
    for i in range(min_length):
        ws[f'A{i+2}'] = coluna_data[i]
        ws[f'B{i+2}'] = coluna_descricao[i]
        ws[f'C{i+2}'] = coluna_valor[i]
        ws[f'D{i+2}'] = coluna_saldo_disponivel[i]
        ws[f'E{i+2}'] = coluna_identificador_transacao[i]
        ws[f'F{i+2}'] = coluna_categoria_conta_contabil[i]
        ws[f'G{i+2}'] = coluna_status_conciliacao[i]
        ws[f'H{i+2}'] = coluna_notas_comentarios[i]

    # Salvar a planilha
    wb.save('Conciliação Bancária.xlsx')

    logging.info("Planilha de conciliação bancária salva")

# Exemplo de uso:
# Abra uma janela de diálogo para selecionar o arquivo PDF
root = tk.Tk()
root.withdraw()  # Esconder a janela principal
pdf_path = filedialog.askopenfilename(title="Selecione o arquivo PDF")
if pdf_path:
    criar_planilha_conciliacao_bancaria(pdf_path)
