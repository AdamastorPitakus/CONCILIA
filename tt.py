import pdfplumber
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
import logging
import re
import csv

# Configurando o logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

# Variável para rastrear o ID das transações
transaction_id = 1

def extract_data_from_pdf(pdf_path):
    logger.info(f"Extraindo dados do PDF: {pdf_path}")
    data = []
    start_extracting = False
    current_date = None
    description_lines = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            page_data = page_text.split('\n')
            
            for line in page_data:
                # Se a linha corresponder ao critério "SALDO EM dd/mm", ative a extração
                if re.match(r'^SALDO EM \d{2}/\d{2}', line):
                    start_extracting = True
                    continue

                # Se a variável start_extracting for True, comece a processar as linhas
                if start_extracting:
                    # Se a linha contém uma data, atualize a data atual e processe qualquer descrição pendente
                    if re.match(r'^\d{2}/\d{2}', line):
                        current_date = re.match(r'^\d{2}/\d{2}', line).group()
                        if description_lines:
                            description = " ".join(description_lines)
                            data.append((current_date, description, None))
                            description_lines = []
                    # Se a linha contém um valor de transação, adicione a data, a descrição e o valor à lista
                    elif re.search(r'- \d+,\d{2}-', line):
                        description = " ".join(description_lines)
                        value = re.search(r'- \d+,\d{2}-', line).group().replace(' ', '').replace('-', '')
                        data.append((current_date, description, value))
                        description_lines = []
                    else:
                        # Se a linha não contiver uma data ou valor, adicione-a à lista de linhas de descrição
                        description_lines.append(line)

    return data

def formatar_valor(valor):
    try:
        is_debit = '-' in valor[-1]
        valor = float(valor.replace('.', '').replace(',', '.').replace('-', ''))
        if is_debit:
            valor = -valor
        return f"R$ {valor:.2f}".replace('.', 'X').replace(',', '.').replace('X', ',')
    except:
        return valor

def criar_planilha_conciliacao_bancaria(pdf_path):
    global transaction_id
    
    logger.info("Criando planilha de conciliação bancária")
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação Bancária"
    # Definição dos cabeçalhos
    headers = ["Identificador de Transação", "Data da Transação", "Descrição/Referência", "Valor", "Saldo Disponível", "Categoria/Conta Contábil", "Status de Conciliação", "Notas/Comentários/NºDoc"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    dados_extraidos = extract_data_from_pdf(pdf_path)
    row_num = 2
    for current_date, description, value in dados_extraidos:
        ws.cell(row=row_num, column=1, value=transaction_id)
        ws.cell(row=row_num, column=2, value=current_date)
        ws.cell(row=row_num, column=3, value=description)
        if value:
            ws.cell(row=row_num, column=4, value=formatar_valor(value))
        transaction_id += 1
        row_num += 1

    wb.save('Conciliação Bancária.xlsx')
    logger.info("Planilha de conciliação bancária salva")
    logger.info(f"Total de dados lidos do PDF: {len(dados_extraidos)}")
    logger.info(f"Total de linhas criadas com ID de operações: {row_num - 2}")

def export_to_csv(data, filename='Concilia.csv'):
    logger.info("Exportando dados para CSV")
    with open(filename, 'w' , newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile)
        #escrever cabeçalhos
        csvwriter.writerow(["Identificador de Transação", "Data da Transação", "Descrição/Referência", "Valor", "Saldo Disponível", "Categoria/Conta Contábil", "Status de Conciliação", "Notas/Comentários/NºDoc"])
        #escrever dados
        for row in data:
            csvwriter.writerow(row)
            logger.info("Dados exportados para CSV")

# Exemplo de uso
root = tk.Tk()
root.withdraw()
pdf_path = filedialog.askopenfilename(title="Selecione o arquivo PDF")
if pdf_path:
    criar_planilha_conciliacao_bancaria(pdf_path)
    export_to_csv(extract_data_from_pdf(pdf_path))
