import pdfplumber
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
import logging
import re

# Configurando o logging para exibir mensagens no terminal e salvar em um arquivo de log com o nome app.log.
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

# Variável para rastrear o ID das transações
transaction_id = 1


def extract_data_from_pdf(pdf_path):
    
    logger.info(f"Extraindo dados do PDF: {pdf_path}")

    data = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            page_data = page_text.split('\n')
            data.extend(page_data)

    for line in data:
        transactions = re.findall(r'(\d{2}/\d{2})\s(.*?)\s(-?\d+,\d{2}-?)\s(-?\d+,\d{2}-?)?', line)
        for transaction in transactions:
            date, description, debit, credit = transaction

    # Salvar os dados extraídos em um arquivo .txt
    with open('dados_extraidos.txt', 'w', encoding='utf-8') as txt_file:
        for line in data:
            txt_file.write(line + '\n')

    logger.info(f"Dados extraídos do PDF: {pdf_path}")
    return data

def criar_planilha_conciliacao_bancaria(pdf_path):
    
    logger.info("Criando planilha de conciliação bancária")

    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação Bancária"
    ws['A1'] = "Data da Transação"
    ws['B1'] = "Descrição/Referência"
    ws['C1'] = "Valor"
    ws['D1'] = "Saldo Disponível"
    ws['E1'] = "Identificador de Transação"
    ws['F1'] = "Categoria/Conta Contábil"
    ws['G1'] = "Status de Conciliação"
    ws['H1'] = "Notas/Comentários"

    logger.info("Planilha de conciliação bancária criada")

    dados_extraidos = extract_data_from_pdf(pdf_path)

    # Inicializar listas vazias para cada coluna
    coluna_data = []
    coluna_descricao_referencia = []
    coluna_valor = []
    coluna_saldo_disponivel = []
    coluna_identificador_transacao = []
    coluna_categoria_conta_contabil = []
    coluna_status_conciliacao = []
    coluna_notas_comentarios = []

    # Inicializar variáveis para acompanhar a coluna atual
    coluna_atual = 0

    # Processar os dados extraídos
    for dado in dados_extraidos:
        if dado.strip():  # Verificar se a linha não está em branco
            # Verificar se a linha contém uma data válida
            if re.match(r'\d{2}/\d{2}', dado):
                coluna_data.append(dado)
                coluna_atual = 0
            # Verificar se a linha contém a descrição/referência
            elif re.match(r'^[^-]', dado):
                coluna_descricao_referencia.append(dado)
                coluna_atual = 1
            # Verificar se a linha contém um valor (crédito ou débito)
            elif re.match(r'-?\d+\.\d{2}', dado):
                coluna_valor.append(float(dado.replace(',', '').replace('-', '')))
                coluna_atual = 2
            # Verificar se a linha contém um saldo disponível
            elif re.match(r'\d+\.\d{2}-', dado):
                coluna_saldo_disponivel.append(float(dado.replace('-', '')))
                coluna_atual = 3
            # Verificar se a linha contém um identificador de transação
            elif re.match(r'I\d+', dado):
                coluna_identificador_transacao.append(dado)
                coluna_atual = 4
            # Verificar se a linha contém uma categoria/conta contábil
            elif re.match(r'C\d+', dado):
                coluna_categoria_conta_contabil.append(dado)
                coluna_atual = 5
            # Verificar se a linha contém um status de conciliação
            elif re.match(r'E\d+', dado):
                coluna_status_conciliacao.append(dado)
                coluna_atual = 6
            # Verificar se a linha contém notas/comentários
            elif re.match(r'N\d+', dado):
                coluna_notas_comentarios.append(dado)
                coluna_atual = 7
            else:
                # Se não se encaixar em nenhuma coluna, inserir na coluna de descrição/referência
                coluna_descricao_referencia[-1] += f" {dado}"

    # Garantir que todas as listas tenham o mesmo comprimento
    min_length = min(len(coluna_data), len(coluna_descricao_referencia), len(coluna_valor),
                     len(coluna_saldo_disponivel), len(coluna_identificador_transacao),
                     len(coluna_categoria_conta_contabil), len(coluna_status_conciliacao),
                     len(coluna_notas_comentarios))

    # Percorrer e gravar os dados nas colunas da planilha
    for i in range(min_length):
        ws[f'A{i+2}'] = coluna_data[i]
        ws[f'B{i+2}'] = coluna_descricao_referencia[i]#.replace('  ', ' ') # Remover espaços extras da descrição da transação (opcional) 
        ws[f'C{i+2}'] = coluna_valor[i] # Adicionando o valor da transação na coluna Valor da planilha de conciliação bancária
        ws[f'D{i+2}'] = coluna_saldo_disponivel[i]# Adicionando o saldo disponível na coluna Saldo Disponível da planilha de conciliação bancária
        ws[f'E{i+2}'] = transaction_id  # Adicionando o ID autoncrementado
        ws[f'F{i+2}'] = coluna_categoria_conta_contabil[i]
        ws[f'G{i+2}'] = coluna_status_conciliacao[i]
        ws[f'H{i+2}'] = coluna_notas_comentarios[i]
        transaction_id += 1  # Incrementando o ID

    # Salvar a planilha
    wb.save('Conciliação Bancária.xlsx')

    logger.info("Planilha de conciliação bancária salva")
    logger.info(f"{min_length} linhas foram importadas e salvas no arquivo.")
    logger.info(f"Total de dados extraídos do PDF: {len(dados_extraidos)}")
    

# Exemplo de uso:
# Abra uma janela de diálogo para selecionar o arquivo PDF
root = tk.Tk()
root.withdraw()  # Esconder a janela principal
pdf_path = filedialog.askopenfilename(title="Selecione o arquivo PDF")
if pdf_path:
    criar_planilha_conciliacao_bancaria(pdf_path)
