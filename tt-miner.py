from pdfminer.high_level import extract_text
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
import logging
import re
import csv
import uuid

# Configurando o logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

def extract_data_from_pdf(pdf_path):
    logger.info(f"Extraindo dados do PDF: {pdf_path}")
    data = []
    current_date = None
    description_lines = []

    pdf_text = extract_text(pdf_path)
    pdf_data = pdf_text.split('\n')

    for line in pdf_data:
        # Se a linha começa com uma data (formato dd/mm), atualizamos a current_date
        if re.match(r'^\d{2}/\d{2}', line):
            current_date = re.match(r'^\d{2}/\d{2}', line).group()
            description_lines = []  # Resetamos a descrição a cada nova data encontrada
            continue

        # Se encontrarmos um valor (identificado pelo padrão - \d+,\d{2}-), adicionamos a current_date,
        # a descrição acumulada e o valor à lista data
        elif re.search(r'- \d+,\d{2}-', line) and current_date:
            description = " ".join(description_lines)
            value = re.search(r'- \d+,\d{2}-', line).group().replace(' ', '').replace('-', '')
            data.append((current_date, description, value))
            description_lines = []  # Resetamos a descrição após identificar um valor
            continue

        # Se a linha não contiver uma data ou valor, a tratamos como parte da descrição
        # e a adicionamos à lista description_lines
        else:
            description_lines.append(line)

    return data

def formatar_valor(valor):
    try:
        is_debit = '-' in valor[-1]
        valor = re.sub(r'[^0-9,.-]', '', valor)
        valor = float(valor.replace(',', '.').replace('-', ''))
        if is_debit:
            valor = -valor
        return f"R$ {valor:.2f}".replace('.', 'X').replace(',', '.').replace('X', ',')
    except:
        return valor

def limpar_texto(texto):
    texto_limpo = re.sub(r'[\/:*?"<>|\[\]\n\r\t♀]', ' ', texto)
    texto_limpo = re.sub(r'\s+', ' ', texto_limpo)
    texto_limpo = texto_limpo.strip()
    return texto_limpo

def criar_planilha_conciliacao_bancaria(dados_extraidos):
    logger.info("Criando planilha de conciliação bancária")
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação Bancária"
    headers = ["Identificador de Transação", "Data da Transação", "Descrição/Referência", "Valor"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    row_num = 2
    for current_date, description, value in dados_extraidos:
        transaction_id = str(uuid.uuid4())
        ws.cell(row=row_num, column=1, value=transaction_id)
        ws.cell(row=row_num, column=2, value=current_date)
        cleaned_description = limpar_texto(description)
        ws.cell(row=row_num, column=3, value=cleaned_description)
        ws.cell(row=row_num, column=4, value=formatar_valor(value))
        row_num += 1

    wb.save('Conciliação Bancária.xlsx')
    logger.info("Planilha de conciliação bancária salva")
    logger.info(f"Total de dados lidos do PDF: {len(dados_extraidos)}")
    logger.info(f"Total de linhas criadas com ID de operações: {row_num - 2}")

def export_to_csv(data, filename='Concilia.csv'):
    logger.info("Exportando dados para CSV")
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile)
        csvwriter.writerow(["Identificador de Transação", "Data da Transação", "Descrição/Referência", "Valor"])
        for row in data:
            csvwriter.writerow([str(uuid.uuid4())] + list(row))
    logger.info("Dados exportados para CSV")

# Exemplo de uso
root = tk.Tk()
root.withdraw()
pdf_path = filedialog.askopenfilename(title="Selecione o arquivo PDF")
if pdf_path:
    extracted_data = extract_data_from_pdf(pdf_path)
    criar_planilha_conciliacao_bancaria(extracted_data)
    export_to_csv(extracted_data)
